"""2-Reiter-Excel-Writer für den CRM-Check-Output.

Reiter 1 "Check": Original-Spalten (20) + 4 zusätzliche:
  aktuell (Ja/Nein/?), bemerkung, konfidenz (0-1), quellen_count

Reiter 2 "Anreicherung": Per-Person-Enrichment:
  row_idx, name, linkedin_url, wikipedia_url, twitter_url, wikidata_id,
  last_press_mention, last_press_title, last_press_url, position_now,
  company_now, address_now, role_change_detected, role_change_note,
  evidence_json (full FieldVerdict-Liste als JSON für Debugging)
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from crm_check.graph.state import CrmCheckState
from crm_check.parser import EXPECTED_HEADER

CHECK_COLUMNS = ["aktuell", "bemerkung", "konfidenz", "tier", "quellen_count"]

ENRICHMENT_COLUMNS = [
    "row_idx",
    "name",
    "verification_tier",
    "score",
    "nor_status",
    "nor_note",
    "linkedin_url",
    "wikipedia_url",
    "twitter_url",
    "wikidata_id",
    "last_press_mention",
    "last_press_title",
    "last_press_url",
    "position_now",
    "company_now",
    "address_now",
    "role_change_detected",
    "role_change_note",
    "evidence_json",
]


def _nor_label(s: str | None) -> str:
    if not s:
        return "?"
    return {"public": "PUBLIC", "nor": "NOR", "unidentified": "UNIDENTIFIED"}.get(s, s.upper())


def _tier_label(t: str | None) -> str:
    if not t:
        return "?"
    return {"confirmed": "Confirmed", "probable": "Probable", "unconfirmed": "Unconfirmed"}.get(t, t.title())


def _aktuell_label(val: bool | None) -> str:
    if val is True:
        return "Ja"
    if val is False:
        return "Nein"
    return "?"


def _aktuell_fill(val: bool | None) -> PatternFill | None:
    if val is True:
        return PatternFill("solid", fgColor="D5E8D4")  # grün
    if val is False:
        return PatternFill("solid", fgColor="F8CECC")  # rot
    return PatternFill("solid", fgColor="E1D5E7")     # lila/grau


def _autosize(ws, widths: dict[int, int]) -> None:
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def write_workbook(
    output_path: Path,
    states: Iterable[CrmCheckState],
) -> Path:
    """Schreibt die 2-Reiter-Excel."""
    wb = Workbook()
    ws_check = wb.active
    ws_check.title = "Check"
    ws_enrich = wb.create_sheet("Anreicherung")

    # Header Reiter 1
    headers = list(EXPECTED_HEADER) + CHECK_COLUMNS
    for col_idx, h in enumerate(headers, start=1):
        cell = ws_check.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DAE8FC")

    # Header Reiter 2
    for col_idx, h in enumerate(ENRICHMENT_COLUMNS, start=1):
        cell = ws_enrich.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D5E8D4")

    aktuell_col_idx = len(EXPECTED_HEADER) + 1  # 21
    check_row = 2
    enrich_row = 2

    for state in states:
        # Reiter 1 — Original-Spalten 1:1
        raw = state.get("raw_row") or {}
        for col_idx, h in enumerate(EXPECTED_HEADER, start=1):
            ws_check.cell(row=check_row, column=col_idx, value=raw.get(h))

        verdict = state.get("verdict")
        enr_state = state.get("enrichment")
        sources_count = 0
        if verdict:
            sources_count = sum(len(fv.sources) for fv in verdict.field_verdicts)
            ws_check.cell(row=check_row, column=aktuell_col_idx,
                          value=_aktuell_label(verdict.aktuell))
            ws_check.cell(row=check_row, column=aktuell_col_idx + 1, value=verdict.bemerkung)
            ws_check.cell(row=check_row, column=aktuell_col_idx + 2,
                          value=round(verdict.konfidenz, 2))
            ws_check.cell(
                row=check_row, column=aktuell_col_idx + 3,
                value=_tier_label(enr_state.verification_tier) if enr_state else "?",
            )
            ws_check.cell(row=check_row, column=aktuell_col_idx + 4, value=sources_count)
            fill = _aktuell_fill(verdict.aktuell)
            if fill:
                ws_check.cell(row=check_row, column=aktuell_col_idx).fill = fill
        else:
            ws_check.cell(row=check_row, column=aktuell_col_idx, value="?")
            ws_check.cell(row=check_row, column=aktuell_col_idx + 1,
                          value="Kein Verdict berechnet.")
            ws_check.cell(row=check_row, column=aktuell_col_idx + 2, value=0)
            ws_check.cell(row=check_row, column=aktuell_col_idx + 3, value="?")

        # Reiter 2 — Anreicherung (Pipeline-v2: tier, score, NOR vorne)
        enr = state.get("enrichment")
        if enr:
            evidence_json = json.dumps(
                [
                    fv.model_dump(mode="json")
                    for fv in (verdict.field_verdicts if verdict else [])
                ],
                ensure_ascii=False,
            )
            ws_enrich.cell(row=enrich_row, column=1, value=state.get("row_idx"))
            ws_enrich.cell(row=enrich_row, column=2, value=state.get("clean_name"))
            ws_enrich.cell(row=enrich_row, column=3, value=_tier_label(enr.verification_tier))
            ws_enrich.cell(row=enrich_row, column=4, value=enr.score)
            ws_enrich.cell(row=enrich_row, column=5, value=_nor_label(enr.nor_status))
            ws_enrich.cell(row=enrich_row, column=6, value=enr.nor_note)
            ws_enrich.cell(row=enrich_row, column=7, value=enr.linkedin_url)
            ws_enrich.cell(row=enrich_row, column=8, value=enr.wikipedia_url)
            ws_enrich.cell(row=enrich_row, column=9, value=enr.twitter_url)
            ws_enrich.cell(row=enrich_row, column=10, value=enr.wikidata_id)
            ws_enrich.cell(
                row=enrich_row, column=11,
                value=enr.last_press_mention.isoformat() if enr.last_press_mention else None,
            )
            ws_enrich.cell(row=enrich_row, column=12, value=enr.last_press_title)
            ws_enrich.cell(row=enrich_row, column=13, value=enr.last_press_url)
            ws_enrich.cell(row=enrich_row, column=14, value=enr.position_now)
            ws_enrich.cell(row=enrich_row, column=15, value=enr.company_now)
            ws_enrich.cell(row=enrich_row, column=16, value=enr.address_now)
            ws_enrich.cell(row=enrich_row, column=17,
                           value="Ja" if enr.role_change_detected else "Nein")
            ws_enrich.cell(row=enrich_row, column=18, value=enr.role_change_note)
            ws_enrich.cell(row=enrich_row, column=19, value=evidence_json)
            enrich_row += 1

        check_row += 1

    # Spaltenbreiten + Wrap
    _autosize(ws_check, {1: 6, 2: 25, 3: 30, 4: 28, 5: 12, 6: 18, 7: 6,
                         aktuell_col_idx: 8, aktuell_col_idx + 1: 60,
                         aktuell_col_idx + 2: 10, aktuell_col_idx + 3: 14,
                         aktuell_col_idx + 4: 12})
    for col in range(aktuell_col_idx + 1, aktuell_col_idx + 2):
        for r in range(2, check_row):
            ws_check.cell(row=r, column=col).alignment = Alignment(wrap_text=True)
    _autosize(ws_enrich, {1: 6, 2: 25, 3: 14, 4: 8, 5: 14, 6: 40,
                          7: 40, 8: 40, 9: 30, 10: 14, 11: 18, 12: 40,
                          13: 40, 14: 30, 15: 25, 16: 25, 17: 8, 18: 30, 19: 60})

    ws_check.freeze_panes = "A2"
    ws_enrich.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


# ── Sonnet-Pivot 2026-06-30 ─────────────────────────────────────────────
# Schreibt CrmContact + SonnetVerdict statt CrmCheckState. Kein LangGraph
# mehr im Datenfluss — Sonnet 4.6 liefert das Verdict direkt.

SONNET_CHECK_COLUMNS = [
    "aktuell", "bemerkung", "konfidenz",
    "neue_position", "neue_firma", "linkedin", "quellen",
    "quellen_count",
]

SONNET_ENRICH_COLUMNS = [
    "row_idx", "name", "position_alt", "firma_alt",
    "aktuell", "konfidenz", "neue_position", "neue_firma",
    "linkedin", "quellen",
]


def write_sonnet_workbook(rows, verdicts, output_path):
    """Schreibt 2-Reiter-Excel aus (CrmContact, SonnetVerdict)-Paaren.

    rows: Sequence[CrmContact] (Reihenfolge = Excel-Reihenfolge)
    verdicts: Sequence[SonnetVerdict] (row_idx-keyed, ggf. luckig)
    """
    from pathlib import Path as _Path
    output_path = _Path(output_path)

    by_idx = {v.row_idx: v for v in verdicts}

    wb = Workbook()
    ws_check = wb.active
    ws_check.title = "Check"
    ws_enrich = wb.create_sheet("Anreicherung")

    headers = list(EXPECTED_HEADER) + SONNET_CHECK_COLUMNS
    for col_idx, h in enumerate(headers, start=1):
        cell = ws_check.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DAE8FC")

    for col_idx, h in enumerate(SONNET_ENRICH_COLUMNS, start=1):
        cell = ws_enrich.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D5E8D4")

    aktuell_col = len(EXPECTED_HEADER) + 1
    check_row = 2
    enrich_row = 2

    for r in rows:
        raw = r.raw or {}
        for col_idx, h in enumerate(EXPECTED_HEADER, start=1):
            ws_check.cell(row=check_row, column=col_idx, value=raw.get(h))

        v = by_idx.get(r.row_idx)
        if v is None:
            ws_check.cell(row=check_row, column=aktuell_col, value="?")
            ws_check.cell(row=check_row, column=aktuell_col + 1, value="Kein Sonnet-Verdict")
            ws_check.cell(row=check_row, column=aktuell_col + 2, value=0.0)
            for off in range(3, len(SONNET_CHECK_COLUMNS)):
                ws_check.cell(row=check_row, column=aktuell_col + off, value="")
        else:
            ws_check.cell(row=check_row, column=aktuell_col + 0, value=_aktuell_label(v.aktuell))
            ws_check.cell(row=check_row, column=aktuell_col + 1, value=v.bemerkung)
            ws_check.cell(row=check_row, column=aktuell_col + 2, value=round(v.konfidenz, 2))
            ws_check.cell(row=check_row, column=aktuell_col + 3, value=v.neue_position or "")
            ws_check.cell(row=check_row, column=aktuell_col + 4, value=v.neue_firma or "")
            ws_check.cell(row=check_row, column=aktuell_col + 5, value=v.linkedin or "")
            ws_check.cell(row=check_row, column=aktuell_col + 6, value="\n".join(v.quellen))
            ws_check.cell(row=check_row, column=aktuell_col + 7, value=len(v.quellen))
            fill = _aktuell_fill(v.aktuell)
            if fill:
                ws_check.cell(row=check_row, column=aktuell_col).fill = fill

            ws_enrich.cell(row=enrich_row, column=1, value=r.row_idx)
            ws_enrich.cell(row=enrich_row, column=2, value=r.name_only)
            ws_enrich.cell(row=enrich_row, column=3, value=r.position)
            ws_enrich.cell(row=enrich_row, column=4, value=r.company)
            ws_enrich.cell(row=enrich_row, column=5, value=_aktuell_label(v.aktuell))
            ws_enrich.cell(row=enrich_row, column=6, value=round(v.konfidenz, 2))
            ws_enrich.cell(row=enrich_row, column=7, value=v.neue_position or "")
            ws_enrich.cell(row=enrich_row, column=8, value=v.neue_firma or "")
            ws_enrich.cell(row=enrich_row, column=9, value=v.linkedin or "")
            ws_enrich.cell(row=enrich_row, column=10, value="\n".join(v.quellen))
            enrich_row += 1

        check_row += 1

    # 20 Original-Spalten + 8 Sonnet-Spalten = 28
    _autosize(ws_check, {
        i: w for i, w in enumerate(
            [6, 6, 14, 10, 8, 18, 8, 28, 28, 22,
             22, 22, 28, 22, 22, 16, 16, 16, 16, 16,
             8, 60, 8, 28, 25, 50, 60, 8],
            start=1,
        )
    })
    # Wrap-Text fuer bemerkung + quellen
    for col_off in (1, 6):
        for rr in range(2, check_row):
            ws_check.cell(row=rr, column=aktuell_col + col_off).alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(ws_enrich, {1: 6, 2: 25, 3: 22, 4: 25, 5: 8, 6: 8, 7: 28, 8: 25, 9: 50, 10: 60})

    ws_check.freeze_panes = "A2"
    ws_enrich.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
